#! python3
#collegeExcel.py
import openpyxl,os
print('ENTER FILENAME')
filename=input()
fileObj=open(filename)
wb=openpyxl.Workbook()
wb.create_sheet('Colleges')
sheet=wb['Colleges']
fileData=fileObj.read()
lines=fileData.split('\n')
for i in range(1,len(lines)):
    print(lines[i])
    words=lines[i].split()
    sheet.cell(row=i+1,column=1).value=' '.join(words[:-4])
    sheet.cell(row=i+1,column=2).value=words[-2]
    sheet.cell(row=i+1,column=3).value=words[-1]
    

wb.save(os.path.join('C:/Users/School Projects/Desktop',os.path.basename(filename)+'.xlsx'))
    
