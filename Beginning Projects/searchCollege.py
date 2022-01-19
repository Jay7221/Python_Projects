#! python3
#searchNITs.py
print('ENTER SEARCH VALUE')
res=input()
print('ENTER FILENAME')
filename=input().lower()
import openpyxl
wb=openpyxl.load_workbook(filename)
sheet=wb['Colleges']
for i in range(1,sheet.max_row+1):
    if res in sheet.cell(row=i,column=1).value.lower():
        print(sheet.cell(row=i,column=1).value.ljust(100),sheet.cell(row=i,column=2).value.rjust(20),sheet.cell(row=i,column=3).value.rjust(20))
