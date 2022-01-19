#! python3
#updateProduce.py - Corrects costs in produce sales spreadesheet.
import openpyxl,os
wb=openpyxl.load_workbook('automate_online-materials\\produceSales.xlsx')
sheet=wb['Sheet']

PRICE_UPDATES={'Garlic':3.07,'Celry':1.19,'Lemon':1.27}

for rowNum in range(1,sheet.max_row + 1):
    produceName=sheet.cell(row=rowNum,column=1).value
    if produceName in PRICE_UPDATES:
        sheet.cell(row=rowNum,column=1).value=PRICE_UPDATES[produceName]
        
    wb.save('automate_online-materials\\updatedProduceSales.xlsx')
    print('Product Updated.')
print('Done!')
