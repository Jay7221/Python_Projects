#! python3
# readCensusExcel.py - Tabulates population and number of census tracts for each county.
import openpyxl,pprint
print('OPENING WORKBOOOK.....')
wb=openpyxl.load_workbook('automate_online-materials\\censuspopdata.xlsx')
sheet=wb['Population by Census Tract']
countyData={}

print('Reading rows....')

for row in range(2,sheet.max_row + 1):
    #Each row in sheet has data for one census tract.
    state=sheet['B'+str(row)].value
    county=sheet['C'+str(row)].value
    population=sheet['D'+str(row)].value

    countyData.setdefault(state,{})
    countyData[state].setdefault(county,{'tracts':0,'population':0})
    countyData[state][county]['population']=countyData[state][county]['population'] + population
    countyData[state][county]['tracts']=countyData[state][county]['tracts'] + 1

pprint.pprint(countyData)


sheet=wb.create_sheet('county data')
sheet.cell(row=1,column=1).value='State'
sheet.cell(row=1,column=2).value='County'
sheet.cell(row=1,column=3).value='Tracts'
sheet.cell(row=1,column=4).value='Poulation'

i=2
for state in countyData:
    for county,data in countyData[state].items():
        sheet.cell(row=i,column=1).value=state
        sheet.cell(row=i,column=2).value=county
        sheet.cell(row=i,column=3).value=data['tracts']
        sheet.cell(row=i,column=4).value=data['population']
        i=i+1
        
            
            
            
            
            
            
wb.save('automate_online-materials\\countyData.xlsx')
print('Done!')
